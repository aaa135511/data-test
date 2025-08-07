import os
import json
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from copy import copy

# --- 配置区 ---
COLUMNS_TO_COMPARE = {
    '计划航班数': '计划航班数',
    '起飞正常率': '起飞正常率',
    '放行正常率': '放行正常率'
}
ORGANIZATION_COLUMN_NAME = '组织中文名'


def parse_to_float(value):
    """
    将不同格式的值（百分比字符串、小数、整数）统一转换为浮点数。
    '98.5%' -> 98.5
    0.985 -> 98.5
    98.5 -> 98.5
    """
    try:
        if isinstance(value, str) and '%' in value:
            return float(value.strip('%'))

        numeric_val = float(value)
        if -1 <= numeric_val <= 1 and numeric_val != 0:
            return numeric_val * 100
        return numeric_val
    except (ValueError, TypeError):
        return None


def main():
    """主执行函数"""
    # 1. 加载环境变量并构建文件名
    load_dotenv()
    date_str = os.getenv('DATE')
    if not date_str:
        print("错误: 未在 .env 文件中找到 'DATE' 配置。")
        return

    input_filename = f'Airport-Normal-{date_str}000000-{int(date_str) + 1}000000.xlsx'
    json_filename = f'competitor_data_{date_str}.json'

    input_filepath = os.path.join('input', input_filename)
    json_filepath = os.path.join('json', json_filename)
    output_dir = 'output'
    output_filepath = os.path.join(output_dir, input_filename)

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs('json', exist_ok=True)

    # 2. 检查文件是否存在
    if not os.path.exists(input_filepath) or not os.path.exists(json_filepath):
        print(f"错误: 确保输入文件和JSON文件都存在。\n- {input_filepath}\n- {json_filepath}")
        return

    # 3. 读取JSON数据
    with open(json_filepath, 'r', encoding='utf-8') as f:
        competitor_list = json.load(f)
    competitor_data_map = {item[ORGANIZATION_COLUMN_NAME]: item for item in competitor_list}
    print(f"成功加载 {len(competitor_data_map)} 条竞品数据。")

    # 4. 加载Excel工作簿和工作表
    workbook = openpyxl.load_workbook(input_filepath)
    try:
        sheet = workbook[workbook.sheetnames[1]]
    except IndexError:
        sheet = workbook.active
    print(f"正在处理工作表: '{sheet.title}'")

    red_font = Font(color="FF0000", name='Calibri', size=11)

    # 5. 查找表头和列索引
    header = [cell.value for cell in sheet[1]]
    try:
        org_col_idx = header.index(ORGANIZATION_COLUMN_NAME) + 1
        col_indices_to_process = [i + 1 for i, h in enumerate(header) if h in COLUMNS_TO_COMPARE]
    except ValueError as e:
        print(f"错误: 在Excel表头中找不到必要的列: {e}")
        return

    # 从右到左插入列
    for col_idx in sorted(col_indices_to_process, reverse=True):
        original_header_name = sheet.cell(row=1, column=col_idx).value

        sheet.insert_cols(col_idx + 1, amount=2)

        header_style = copy(sheet.cell(row=1, column=col_idx)._style)

        competitor_header_cell = sheet.cell(row=1, column=col_idx + 1)
        competitor_header_cell.value = f'飞常准{original_header_name}'
        competitor_header_cell._style = header_style

        diff_header_cell = sheet.cell(row=1, column=col_idx + 2)
        diff_header_cell.value = '差值'
        diff_header_cell._style = header_style

        print(f"已在 '{original_header_name}' ({get_column_letter(col_idx)}) 右侧插入两列。")

        # 6. 遍历行，填充数据和差值
        for row in range(2, sheet.max_row + 1):
            org_name = sheet.cell(row=row, column=org_col_idx).value

            if org_name and org_name in competitor_data_map:
                competitor_item = competitor_data_map[org_name]
                json_key = COLUMNS_TO_COMPARE.get(original_header_name)

                if json_key and json_key in competitor_item:
                    our_cell = sheet.cell(row=row, column=col_idx)
                    competitor_value = competitor_item[json_key]

                    base_style = copy(our_cell._style)
                    competitor_cell = sheet.cell(row=row, column=col_idx + 1)
                    diff_cell = sheet.cell(row=row, column=col_idx + 2)

                    # --- 新增逻辑: 专门处理计划航班数为 "N/A" 的情况 ---
                    if original_header_name == '计划航班数' and str(competitor_value).strip().upper() == 'N/A':
                        # 直接填充 "N/A" 并设置样式
                        competitor_cell.value = "N/A"
                        competitor_cell._style = base_style
                        competitor_cell.font = copy(red_font)

                        diff_cell.value = "N/A"
                        diff_cell._style = base_style
                        diff_cell.font = copy(red_font)
                    else:
                        # --- 保持原有计算逻辑 ---
                        our_value = our_cell.value

                        # 写入竞品数据
                        competitor_cell._style = base_style
                        comp_float = parse_to_float(competitor_value)
                        if comp_float is not None:
                            if "正常率" in original_header_name:
                                competitor_cell.value = comp_float / 100
                                competitor_cell.number_format = '0.00%'
                            else:
                                competitor_cell.value = comp_float
                        else:
                            competitor_cell.value = competitor_value
                        competitor_cell.font = copy(red_font)

                        # 计算并写入差值
                        diff_cell._style = base_style
                        our_float = parse_to_float(our_value)
                        if our_float is not None and comp_float is not None:
                            diff = our_float - comp_float
                            if "正常率" in original_header_name:
                                diff_cell.value = diff / 100
                                diff_cell.number_format = '0.00%'
                            else:
                                diff_cell.value = diff
                        else:
                            diff_cell.value = "N/A"
                        diff_cell.font = copy(red_font)

    # 7. 保存工作簿
    try:
        workbook.save(output_filepath)
        print(f"\n处理完成！\n结果已保存至: {output_filepath}")
    except PermissionError:
        print(f"\n错误: 无法保存文件。请确保文件 '{output_filepath}' 没有被其他程序打开。")


if __name__ == '__main__':
    main()