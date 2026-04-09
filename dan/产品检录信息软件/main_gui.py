import sys
import os
import sqlite3
import io
import re
import zipfile
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLineEdit, QPushButton, QTreeWidget, QTreeWidgetItem, QLabel,
    QFileDialog, QScrollArea, QFrame, QSplitter, QProgressBar, QMessageBox,
    QTextEdit, QGridLayout, QCheckBox, QDialog
)
from PyQt6.QtGui import QPixmap, QColor, QFont, QCursor
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSettings
from openpyxl import load_workbook
from lxml import etree

# --- 全局配置 ---
DB_PATH = "products_pro_v3.db"
MARKUP_FACTOR = 1.2  # 普通用户价格倍数


# --- 登录对话框 (支持记住密码 & 最小化) ---
class LoginDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("登录 - LangerMoto 产品检录系统")
        self.setFixedSize(380, 260)
        # 增加最小化按钮支持
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinimizeButtonHint)
        self.settings = QSettings("LangerMotoApp", "LoginSettings")
        self.role = None
        self.init_ui()
        self.load_saved_credentials()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 25, 40, 25)
        layout.setSpacing(12)

        title = QLabel("系统身份验证")
        title.setFont(QFont("Microsoft YaHei", 15, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        self.user_input = QLineEdit()
        self.user_input.setPlaceholderText("请输入账号")
        self.user_input.setFixedHeight(35)
        layout.addWidget(self.user_input)

        self.pwd_input = QLineEdit()
        self.pwd_input.setPlaceholderText("请输入密码")
        self.pwd_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.pwd_input.setFixedHeight(35)
        layout.addWidget(self.pwd_input)

        self.remember_cb = QCheckBox("记住账号密码")
        layout.addWidget(self.remember_cb)

        self.login_btn = QPushButton("登 录")
        self.login_btn.setFixedHeight(40)
        self.login_btn.setStyleSheet("""
            QPushButton { background-color: #1890ff; color: white; font-weight: bold; border-radius: 4px; font-size: 14px; }
            QPushButton:hover { background-color: #40a9ff; }
        """)
        self.login_btn.clicked.connect(self.check_login)
        layout.addWidget(self.login_btn)

    def load_saved_credentials(self):
        """加载保存的账号密码"""
        user = self.settings.value("username", "")
        pwd = self.settings.value("password", "")
        remember = self.settings.value("remember", "false") == "true"

        if remember:
            self.user_input.setText(user)
            self.pwd_input.setText(pwd)
            self.remember_cb.setChecked(True)

    def check_login(self):
        u = self.user_input.text().strip()
        p = self.pwd_input.text().strip()

        # 更新后的账号密码
        is_admin = (u == "LangerMoto" and p == "LangerMoto888")
        is_user = (u == "User" and p == "123321123")

        if is_admin or is_user:
            self.role = "admin" if is_admin else "user"

            # 处理记住密码逻辑
            if self.remember_cb.isChecked():
                self.settings.setValue("username", u)
                self.settings.setValue("password", p)
                self.settings.setValue("remember", "true")
            else:
                self.settings.remove("username")
                self.settings.remove("password")
                self.settings.setValue("remember", "false")

            self.accept()
        else:
            QMessageBox.warning(self, "登录失败", "账号或密码错误，请核对后输入。")


# --- 数据库管理 ---
class DatabaseManager:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        self._init_db()

    def _get_conn(self):
        conn = sqlite3.connect(self.db_path, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self):
        conn = self._get_conn()
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                brand TEXT, model TEXT, name TEXT, oe_no TEXT, sku TEXT, 
                desc_zh TEXT, desc_en TEXT, price TEXT, link TEXT, 
                size TEXT, inventory TEXT, records TEXT, supplier TEXT
            )''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS product_images (
                id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER,
                image_blob BLOB, FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )''')
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_sku ON products(sku)")
        conn.commit();
        conn.close()

    def execute_query(self, sql, params=()):
        conn = self._get_conn();
        cursor = conn.cursor()
        cursor.execute(sql, params);
        res = cursor.fetchall()
        conn.commit();
        conn.close()
        return res


# --- 高清原图查看器 ---
class BigImageViewer(QDialog):
    def __init__(self, image_blob, p_name, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"高清原图 - {p_name}")
        self.setWindowFlags(
            self.windowFlags() | Qt.WindowType.WindowMaximizeButtonHint | Qt.WindowType.WindowMinimizeButtonHint)
        self.resize(1000, 800)
        layout = QVBoxLayout(self)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background-color: #333;")
        self.label = QLabel()
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pix = QPixmap();
        pix.loadFromData(image_blob)
        self.label.setPixmap(pix)  # 显示1:1原图
        scroll.setWidget(self.label)
        layout.addWidget(scroll)


# --- 核心导入线程 (WPS嵌入图兼容) ---
class ImportThread(QThread):
    progress = pyqtSignal(int)
    log = pyqtSignal(str)
    finished = pyqtSignal(int)
    error = pyqtSignal(str)

    def __init__(self, file_path):
        super().__init__();
        self.file_path = file_path

    def run(self):
        try:
            self.log.emit("🔍 正在启动深度解析引擎...")
            wps_map = self.extract_wps_images(self.file_path)
            wb_img = load_workbook(self.file_path, data_only=True)
            ws_img = wb_img.active
            anchor_map = {}
            if hasattr(ws_img, '_images'):
                for img in ws_img._images:
                    try:
                        row = img.anchor._from.row + 1; anchor_map[row] = img._data()
                    except:
                        continue

            wb_data = load_workbook(self.file_path, data_only=False);
            ws_data = wb_data.active
            rows = list(ws_data.iter_rows(min_row=2))

            conn = sqlite3.connect(DB_PATH);
            cursor = conn.cursor()
            count = 0
            for i, row in enumerate(rows):
                sku = str(row[4].value).strip() if row[4].value else None
                if not sku: continue
                price_val = str(row[8].value) if row[8].value is not None else ""
                if "DISPIMG" in price_val: price_val = "0"

                cursor.execute('''INSERT INTO products (brand, model, name, oe_no, sku, desc_zh, desc_en, price, link, size, inventory, records, supplier)
                                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                               (row[0].value, row[1].value, row[2].value, row[3].value, sku, row[5].value, row[6].value,
                                price_val, row[9].value, row[10].value, row[12].value, row[13].value, row[14].value))
                pid = cursor.lastrowid
                blob = None
                for cell in row:
                    if cell.value and "DISPIMG" in str(cell.value):
                        mid = re.search(r'DISPIMG\("(.+?)"', str(cell.value))
                        if mid and mid.group(1) in wps_map: blob = wps_map[mid.group(1)]; break
                if not blob and (i + 2) in anchor_map: blob = anchor_map[i + 2]
                if blob: cursor.execute("INSERT INTO product_images (product_id, image_blob) VALUES (?, ?)",
                                        (pid, sqlite3.Binary(blob)))
                count += 1
                if i % 10 == 0: self.progress.emit(int((i + 1) / len(rows) * 100))
            conn.commit();
            conn.close();
            self.finished.emit(count)
        except Exception as e:
            self.error.emit(f"❌ 导入崩溃: {str(e)}")

    def extract_wps_images(self, path):
        img_map = {}
        try:
            with zipfile.ZipFile(path, 'r') as z:
                rels = etree.fromstring(z.read('xl/_rels/cellimages.xml.rels'))
                rid_to_path = {r.get('Id'): r.get('Target') for r in rels.xpath('//*[local-name()="Relationship"]')}
                xml = etree.fromstring(z.read('xl/cellimages.xml'))
                for entry in xml.xpath('//*[local-name()="cellImage"]'):
                    nid = entry.xpath('.//*[local-name()="cNvPr"]/@name')[0]
                    rid = entry.xpath('.//*[local-name()="blip"]/@*[local-name()="embed"]')[0]
                    p = f"xl/{rid_to_path[rid]}"
                    if p in z.namelist(): img_map[nid] = z.read(p)
        except:
            pass
        return img_map


# --- 主界面 ---
class MainWindow(QMainWindow):
    def __init__(self, role):
        super().__init__()
        self.db = DatabaseManager()
        self.user_role = role  # admin / user
        self.current_product_id = None
        self.current_blobs = []
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f"LangerMoto 产品检录系统 Pro [{'管理员' if self.user_role == 'admin' else '普通用户'}]")
        self.resize(1450, 950)
        self.setStyleSheet("""
            QMainWindow { background-color: #f0f2f5; }
            QLineEdit, QTextEdit { border: 1px solid #ccc; border-radius: 2px; padding: 5px; background: white; }
            QPushButton#SaveBtn { background-color: #1890ff; color: white; font-weight: bold; border-radius: 4px; }
            QPushButton#DelBtn { background-color: #ff4d4f; color: white; border-radius: 4px; }
            QLineEdit:disabled, QTextEdit:disabled { background-color: #f5f5f5; color: #888; }
        """)

        central = QWidget();
        self.setCentralWidget(central);
        main_layout = QVBoxLayout(central)

        # 顶部
        top = QHBoxLayout()
        self.search_in = QLineEdit();
        self.search_in.setPlaceholderText("🔍 快速检索：品名、货号、OE、描述、车型...")
        self.search_in.textChanged.connect(self.do_search);
        top.addWidget(self.search_in, 6)

        self.btn_add = QPushButton("新增");
        self.btn_add.clicked.connect(self.add_new)
        self.btn_save = QPushButton("保存修改");
        self.btn_save.setObjectName("SaveBtn");
        self.btn_save.clicked.connect(self.save_data)
        self.btn_del = QPushButton("删除");
        self.btn_del.setObjectName("DelBtn");
        self.btn_del.clicked.connect(self.del_data)
        self.btn_imp = QPushButton("批量导入");
        self.btn_imp.clicked.connect(self.imp_data)
        self.btn_exp = QPushButton("导出勾选");
        self.btn_exp.clicked.connect(self.exp_data)

        for b in [self.btn_add, self.btn_save, self.btn_del, self.btn_imp, self.btn_exp]:
            b.setFixedSize(90, 35);
            top.addWidget(b)

        if self.user_role == "user":
            self.btn_add.setEnabled(False);
            self.btn_save.setEnabled(False)
            self.btn_del.setEnabled(False);
            self.btn_imp.setEnabled(False)

        main_layout.addLayout(top)
        self.pbar = QProgressBar();
        self.pbar.setFixedHeight(4);
        self.pbar.hide();
        main_layout.addWidget(self.pbar)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        self.tree = QTreeWidget();
        self.tree.setHeaderLabels(["产品目录结构 (勾选导出)"]);
        self.tree.itemClicked.connect(self.on_tree_click);
        self.tree.itemChanged.connect(self.on_tree_changed)
        splitter.addWidget(self.tree)

        right_splitter = QSplitter(Qt.Orientation.Vertical)
        self.scroll = QScrollArea();
        self.scroll.setWidgetResizable(True);
        self.edit_container = QWidget();
        self.grid = QGridLayout(self.edit_container)
        self.setup_edit_fields();
        self.scroll.setWidget(self.edit_container)
        self.log_area = QTextEdit();
        self.log_area.setReadOnly(True);
        self.log_area.setStyleSheet("background:#f9f9f9; font-size:11px;")
        right_splitter.addWidget(self.scroll);
        right_splitter.addWidget(self.log_area)
        right_splitter.setSizes([750, 150]);
        splitter.addWidget(right_splitter)
        splitter.setSizes([350, 1100]);
        main_layout.addWidget(splitter)
        self.do_search()

    def setup_edit_fields(self):
        self.img_label = QLabel("🖼️ 预览图");
        self.img_label.setFixedSize(240, 240);
        self.img_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.img_label.setCursor(QCursor(Qt.CursorShape.PointingHandCursor));
        self.img_label.mousePressEvent = self.show_big_image
        self.img_label.setStyleSheet("border: 1px dashed #aaa; background: #fff;");
        self.grid.addWidget(self.img_label, 0, 0, 4, 2)
        self.btn_gallery = QPushButton("📂 多图库 (0)");
        self.btn_gallery.clicked.connect(self.open_gallery);
        self.grid.addWidget(self.btn_gallery, 4, 0, 1, 2)

        self.edit_brand = QLineEdit();
        self.edit_model = QLineEdit();
        self.edit_name = QLineEdit()
        self.edit_price = QLineEdit();
        self.edit_inv = QLineEdit();
        self.edit_sku = QLineEdit()
        self.edit_oe = QLineEdit();
        self.edit_link = QLineEdit()

        self.grid.addWidget(QLabel("品牌:"), 0, 2);
        self.grid.addWidget(self.edit_brand, 0, 3)
        self.grid.addWidget(QLabel("车型:"), 0, 4);
        self.grid.addWidget(self.edit_model, 0, 5)
        self.grid.addWidget(QLabel("产品名称:"), 1, 2);
        self.grid.addWidget(self.edit_name, 1, 3, 1, 3)
        self.grid.addWidget(QLabel("报价价格:"), 2, 2);
        self.grid.addWidget(self.edit_price, 2, 3)
        self.grid.addWidget(QLabel("库存数量:"), 2, 4);
        self.grid.addWidget(self.edit_inv, 2, 5)
        self.grid.addWidget(QLabel("产品货号:"), 3, 2);
        self.grid.addWidget(self.edit_sku, 3, 3)
        self.grid.addWidget(QLabel("OE号:"), 3, 4);
        self.grid.addWidget(self.edit_oe, 3, 5)
        self.grid.addWidget(QLabel("相关链接:"), 4, 2);
        self.grid.addWidget(self.edit_link, 4, 3, 1, 3)

        self.grid.addWidget(QLabel("🔹 描述与规格"), 5, 0, 1, 6)
        self.edit_zh = QTextEdit();
        self.edit_zh.setFixedHeight(120);
        self.edit_en = QTextEdit();
        self.edit_en.setFixedHeight(120)
        self.grid.addWidget(QLabel("中文描述:"), 6, 0);
        self.grid.addWidget(self.edit_zh, 6, 1, 1, 2)
        self.grid.addWidget(QLabel("英文描述:"), 6, 3);
        self.grid.addWidget(self.edit_en, 6, 4, 1, 2)
        self.edit_specs = QLineEdit();
        self.edit_sup = QLineEdit()
        self.grid.addWidget(QLabel("尺寸重量:"), 7, 0);
        self.grid.addWidget(self.edit_specs, 7, 1, 1, 2)
        self.grid.addWidget(QLabel("供应商:"), 7, 3);
        self.grid.addWidget(self.edit_sup, 7, 4, 1, 2)
        self.edit_rec = QTextEdit();
        self.edit_rec.setFixedHeight(80)
        self.grid.addWidget(QLabel("🔹 内部记录"), 8, 0, 1, 6);
        self.grid.addWidget(self.edit_rec, 9, 1, 1, 5)

        if self.user_role == "user":
            for w in [self.edit_brand, self.edit_model, self.edit_name, self.edit_price, self.edit_inv, self.edit_sku,
                      self.edit_oe, self.edit_link, self.edit_zh, self.edit_en, self.edit_specs, self.edit_sup,
                      self.edit_rec]:
                w.setEnabled(False)

    def write_log(self, text):
        self.log_area.append(f"[{datetime.now().strftime('%H:%M:%S')}] {text}")

    def show_big_image(self, event):
        if self.current_blobs: BigImageViewer(self.current_blobs[0], self.edit_name.text(), self).exec()

    def do_search(self):
        txt = self.search_in.text().strip();
        sql = "SELECT id, brand, model, name, sku FROM products WHERE 1=1";
        params = []
        if txt:
            sql += " AND (name LIKE ? OR sku LIKE ? OR oe_no LIKE ? OR brand LIKE ? OR model LIKE ? OR desc_zh LIKE ? OR desc_en LIKE ?)"
            p = f"%{txt}%";
            params = [p, p, p, p, p, p, p]
        rows = self.db.execute_query(sql, params);
        self.update_tree(rows)

    def update_tree(self, data):
        self.tree.blockSignals(True);
        self.tree.clear();
        d_map = {}
        for r in data:
            b, m = r['brand'] or "未分类", r['model'] or "通用"
            if b not in d_map: d_map[b] = {}
            if m not in d_map[b]: d_map[b][m] = []
            d_map[b][m].append((f"{r['name']} ({r['sku']})", r['id']))
        for b in sorted(d_map.keys()):
            bi = QTreeWidgetItem(self.tree, [b]);
            bi.setCheckState(0, Qt.CheckState.Unchecked);
            bi.setFlags(bi.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            for m in sorted(d_map[b].keys()):
                mi = QTreeWidgetItem(bi, [m]);
                mi.setCheckState(0, Qt.CheckState.Unchecked);
                mi.setFlags(mi.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                for n, pid in d_map[b][m]:
                    pi = QTreeWidgetItem(mi, [n]);
                    pi.setCheckState(0, Qt.CheckState.Unchecked);
                    pi.setFlags(pi.flags() | Qt.ItemFlag.ItemIsUserCheckable);
                    pi.setData(0, Qt.ItemDataRole.UserRole, pid)
        self.tree.blockSignals(False);
        if self.search_in.text(): self.tree.expandAll()

    def on_tree_changed(self, item, col):
        self.tree.blockSignals(True);
        state = item.checkState(0)

        def check_all(it, st):
            for i in range(it.childCount()): child = it.child(i); child.setCheckState(0, st); check_all(child, st)

        check_all(item, state);
        self.tree.blockSignals(False)

    def on_tree_click(self, item, col):
        pid = item.data(0, Qt.ItemDataRole.UserRole)
        if pid: self.load_data(pid)

    def load_data(self, pid):
        res = self.db.execute_query("SELECT * FROM products WHERE id=?", (pid,))
        if not res: return
        p = res[0];
        self.current_product_id = pid
        self.edit_brand.setText(p['brand'] or "");
        self.edit_model.setText(p['model'] or "");
        self.edit_name.setText(p['name'] or "")
        self.edit_sku.setText(p['sku'] or "");
        self.edit_oe.setText(p['oe_no'] or "");
        self.edit_zh.setText(p['desc_zh'] or "")
        self.edit_en.setText(p['desc_en'] or "");
        self.edit_specs.setText(p['size'] or "");
        self.edit_inv.setText(p['inventory'] or "")

        if self.user_role == "admin":
            self.edit_price.setText(p['price'] or "");
            self.edit_sup.setText(p['supplier'] or "");
            self.edit_rec.setText(p['records'] or "");
            self.edit_link.setText(p['link'] or "")
        else:
            try:
                price = float(p['price']);
                self.edit_price.setText(str(round(price * MARKUP_FACTOR, 2)))
            except:
                self.edit_price.setText(p['price'] or "面议")
            self.edit_sup.setText("******");
            self.edit_rec.setText("******");
            self.edit_link.setText("******")

        imgs = self.db.execute_query("SELECT image_blob FROM product_images WHERE product_id=?", (pid,))
        self.current_blobs = [r['image_blob'] for r in imgs]
        self.btn_gallery.setText(f"📂 多图库 ({len(self.current_blobs)})")
        if self.current_blobs:
            px = QPixmap();
            px.loadFromData(self.current_blobs[0])
            self.img_label.setPixmap(
                px.scaled(235, 235, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        else:
            self.img_label.setText("🖼️ 暂无图片")

    def save_data(self):
        if self.user_role != "admin" or not self.edit_name.text().strip(): return
        d = (self.edit_brand.text(), self.edit_model.text(), self.edit_name.text(), self.edit_oe.text(),
             self.edit_sku.text(), self.edit_zh.toPlainText(), self.edit_en.toPlainText(), self.edit_price.text(),
             self.edit_link.text(), self.edit_specs.text(), self.edit_inv.text(), self.edit_rec.toPlainText(),
             self.edit_sup.text())
        if self.current_product_id:
            self.db.execute_query(
                "UPDATE products SET brand=?,model=?,name=?,oe_no=?,sku=?,desc_zh=?,desc_en=?,price=?,link=?,size=?,inventory=?,records=?,supplier=? WHERE id=?",
                d + (self.current_product_id,))
            pid = self.current_product_id
        else:
            self.db.execute_query(
                "INSERT INTO products(brand,model,name,oe_no,sku,desc_zh,desc_en,price,link,size,inventory,records,supplier) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                d)
            pid = self.db.execute_query("SELECT last_insert_rowid() as id")[0]['id']
        self.db.execute_query("DELETE FROM product_images WHERE product_id=?", (pid,))
        for b in self.current_blobs: self.db.execute_query(
            "INSERT INTO product_images(product_id,image_blob) VALUES(?,?)", (pid, sqlite3.Binary(b)))
        self.write_log(f"✅ 已更新：{self.edit_name.text()}");
        self.do_search()

    def imp_data(self):
        p, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "Excel (*.xlsx)")
        if p:
            self.log_area.clear();
            self.pbar.show();
            self.task = ImportThread(p)
            self.task.progress.connect(self.pbar.setValue);
            self.task.log.connect(self.write_log)
            self.task.error.connect(lambda m: (self.pbar.hide(), QMessageBox.critical(self, "失败", m)))
            self.task.finished.connect(
                lambda n: (self.pbar.hide(), self.do_search(), QMessageBox.information(self, "成功", f"录入{n}条。")))
            self.task.start()

    def exp_data(self):
        ids = []

        def get_checked(it):
            pid = it.data(0, Qt.ItemDataRole.UserRole)
            if pid and it.checkState(0) == Qt.CheckState.Checked: ids.append(pid)
            for i in range(it.childCount()): get_checked(it.child(i))

        for i in range(self.tree.topLevelItemCount()): get_checked(self.tree.topLevelItem(i))
        if not ids: return
        p, _ = QFileDialog.getSaveFileName(self, "导出", "导出.xlsx", "Excel (*.xlsx)")
        if p:
            from openpyxl import Workbook
            wb = Workbook();
            ws = wb.active;
            ws.append(["品牌", "车型", "名称", "OE", "货号", "中描述", "英描述", "价格", "链接", "尺寸", "库存", "记录",
                       "供应商"])
            rows = self.db.execute_query(
                f"SELECT brand,model,name,oe_no,sku,desc_zh,desc_en,price,link,size,inventory,records,supplier FROM products WHERE id IN ({','.join(['?'] * len(ids))})",
                tuple(ids))
            for r in rows: ws.append(list(r))
            wb.save(p);
            self.write_log("📤 导出完成")

    def add_new(self):
        self.current_product_id = None;
        self.current_blobs = []
        for w in self.edit_container.findChildren((QLineEdit, QTextEdit)): w.clear()
        self.img_label.setText("预览图");
        self.btn_gallery.setText("📂 多图库 (0)")

    def del_data(self):
        if self.current_product_id and QMessageBox.question(self, "?", "删除？") == QMessageBox.StandardButton.Yes:
            self.db.execute_query("DELETE FROM products WHERE id=?", (self.current_product_id,));
            self.add_new();
            self.do_search()

    def open_gallery(self):
        if self.user_role != "admin": return
        dlg = ImageGalleryDialog([{'image_blob': b} for b in self.current_blobs], self)
        if dlg.exec():
            self.current_blobs = dlg.all_blobs;
            self.btn_gallery.setText(f"📂 多图库 ({len(self.current_blobs)})")
            if self.current_blobs:
                px = QPixmap();
                px.loadFromData(self.current_blobs[0])
                self.img_label.setPixmap(
                    px.scaled(235, 235, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))


class ImageGalleryDialog(QDialog):
    def __init__(self, images, parent=None):
        super().__init__(parent);
        self.setWindowTitle("多图管理");
        self.resize(700, 500)
        self.all_blobs = [img['image_blob'] for img in images] if images else []
        layout = QVBoxLayout(self);
        btn = QPushButton("➕ 添加");
        btn.clicked.connect(self.add);
        layout.addWidget(btn)
        self.scroll = QScrollArea();
        self.container = QWidget();
        self.grid = QGridLayout(self.container)
        self.scroll.setWidgetResizable(True);
        self.scroll.setWidget(self.container);
        layout.addWidget(self.scroll)
        b2 = QPushButton("应用");
        b2.clicked.connect(self.accept);
        layout.addWidget(b2);
        self.refresh()

    def add(self):
        ps, _ = QFileDialog.getOpenFileNames(self, "选图", "", "Img (*.png *.jpg *.jpeg)")
        for p in ps:
            with open(p, 'rb') as f: self.all_blobs.append(f.read())
        self.refresh()

    def refresh(self):
        while self.grid.count():
            w = self.grid.takeAt(0).widget();
            if w: w.deleteLater()
        for i, b in enumerate(self.all_blobs):
            l = QLabel();
            px = QPixmap();
            px.loadFromData(b);
            l.setPixmap(
                px.scaled(150, 150, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            db = QPushButton("删");
            db.clicked.connect(lambda ch, idx=i: (self.all_blobs.pop(idx), self.refresh()))
            v = QVBoxLayout();
            v.addWidget(l);
            v.addWidget(db);
            w = QWidget();
            w.setLayout(v);
            self.grid.addWidget(w, i // 3, i % 3)


if __name__ == "__main__":
    app = QApplication(sys.argv);
    app.setStyle("Fusion")
    login = LoginDialog()
    if login.exec() == QDialog.DialogCode.Accepted:
        window = MainWindow(login.role);
        window.show();
        sys.exit(app.exec())
    else:
        sys.exit(0)