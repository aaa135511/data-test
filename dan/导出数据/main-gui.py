import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
import os
import sys
import time
import requests  # 需要安装: pip install requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font


class OrderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("门业自动拆单系统 - 增强网络稳定性版")
        self.root.geometry("850x650")

        icon_path = os.path.join(os.path.dirname(__file__), "app.ico")
        if os.path.exists(icon_path) and sys.platform.startswith('win'):
            self.root.iconbitmap(icon_path)

        self.source_data = []
        self.order_info = {"工程名称": "", "单号": "", "日期": ""}

        # 模拟API配置 (这里你可以替换成真实的接口地址)
        self.api_url = "https://api.example.com/get_price"

        self.setup_ui()

    def setup_ui(self):
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="1. 导入订单 Excel", command=self.import_excel, width=18, bg="#e1e1e1").grid(row=0,
                                                                                                               column=0,
                                                                                                               padx=5)
        tk.Button(btn_frame, text="2. 导出《制作单》", command=lambda: self.export_action("制作单"), width=18).grid(row=0,
                                                                                                                  column=1,
                                                                                                                  padx=5)
        tk.Button(btn_frame, text="3. 导出《出库单》", command=lambda: self.export_action("出库单"), width=18).grid(row=0,
                                                                                                                  column=2,
                                                                                                                  padx=5)
        tk.Button(btn_frame, text="4. 导出《配件单》", command=lambda: self.export_action("配件单"), width=18).grid(row=0,
                                                                                                                  column=3,
                                                                                                                  padx=5)

        self.log_box = scrolledtext.ScrolledText(self.root, height=30, width=100, font=("Courier New", 10))
        self.log_box.pack(padx=20, pady=5)
        self.log("系统启动。如果导出过程中涉及联网数据，程序将具备断网自动重试功能。")

    def log(self, message, color="black"):
        time_str = datetime.now().strftime("%H:%M:%S")
        self.log_box.insert(tk.END, f"[{time_str}] {message}\n")
        # 简单的染色逻辑 (可选)
        if "错误" in message or "失败" in message:
            self.log_box.tag_add("err", "end-2l", "end-1l")
            self.log_box.tag_config("err", foreground="red")
        self.log_box.see(tk.END)
        self.root.update()  # 强制刷新界面，防止等待时卡死

    # --- 核心逻辑：带断网检测的请求函数 ---
    def request_data_safe(self, item_name):
        """
        item_name: 需要查询的项名
        逻辑：如果是网络问题，死循环等待；如果是业务查不到，直接返回None。
        """
        while True:
            try:
                # 这里模拟一个联网请求，设置3秒超时
                # 实际使用时替换为：response = requests.get(self.api_url, params={'name': item_name}, timeout=3)
                # ---------------- 模拟请求开始 ----------------
                # 假设我们是在查询单价。为了测试，我们模拟一下：
                self.log(f"正在联网获取 [{item_name}] 的补充数据...")

                # 模拟逻辑：如果item_name是"不存在"，模拟404；如果是网络断开，抛出异常
                # 实际生产中 requests 会根据网络状态自动抛出 ConnectionError
                response = requests.get("http://www.baidu.com", timeout=3)  # 用百度测试连通性

                # 假设这是你的业务接口逻辑
                # if response.status_code == 404:
                #     self.log(f"⚠️ 查询结果：库中无 [{item_name}] 的数据，已确认，不重试。", "orange")
                #     return "无数据"

                # 模拟成功返回
                return "OK"
                # ---------------- 模拟请求结束 ----------------

            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
                self.log(f"❌ 网络连接中断！请求 [{item_name}] 失败。程序已挂起，正在等待网络恢复...", "red")
                time.sleep(5)  # 每5秒检查一次网络
                continue  # 继续死循环，直到成功获取响应
            except Exception as e:
                self.log(f"程序内部异常: {e}")
                return None

    def import_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path: return
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            self.order_info['工程名称'] = ws['B2'].value
            self.order_info['日期'] = ws['L2'].value
            self.order_info['单号'] = ws['P2'].value

            self.source_data = []
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
                if row[0] is None: continue
                self.source_data.append({
                    "序号": row[0], "级别": row[1], "厚度": row[2], "造型": row[3], "颜色": row[4],
                    "框料": row[5], "扇料": row[6], "总高": row[7], "宽": row[8], "门高": row[9],
                    "亮子": row[10], "填充物": row[11], "开向": row[12], "锁向": row[13],
                    "扣边": row[14], "底坎": row[15], "锁具": row[16], "闭门器": row[17],
                    "窗口": row[18], "数量": row[19], "备注": row[20]
                })
            self.log(f"导入成功。单号：{self.order_info['单号']}")
        except Exception as e:
            self.log(f"导入失败: {e}")

    def export_action(self, target_type):
        if not self.source_data:
            messagebox.showwarning("提示", "请先导入数据！")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 initialfile=f"{target_type}_{self.order_info['单号']}.xlsx")
        if not save_path: return

        try:
            wb = Workbook()
            ws = wb.active
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            align = Alignment(horizontal='center', vertical='center')

            # 在处理每一行数据前，先通过网络校验/获取数据
            self.log(f"开始生成{target_type}，正在执行数据预处理...")
            for item in self.source_data:
                # 比如根据“锁具”型号去网上查单价或验证型号
                status = self.request_data_safe(item['锁具'])
                if status == "无数据":
                    self.log(f"序号 {item['序号']}：锁具型号 [{item['锁具']}] 未在库中找到，跳过增量处理。")

            # 根据类型绘制表格 (逻辑同前，已略去重复的绘制代码，保持简洁)
            if target_type == "制作单":
                self.draw_production_sheet(ws, border, align)
            elif target_type == "出库单":
                self.draw_outbound_sheet(ws, border, align)
            else:
                self.draw_parts_sheet(ws, border, align)

            wb.save(save_path)
            self.log(f"✅ 《{target_type}》导出成功。")
            messagebox.showinfo("成功", "导出成功！")
        except Exception as e:
            self.log(f"导出失败: {e}")

    # --- 以下绘制函数保持之前的精美排版逻辑 ---
    def draw_production_sheet(self, ws, border, align):
        ws.title = "制作单"
        ws.merge_cells('A1:U1')
        ws['A1'] = "制 作 单"
        ws['A1'].font = Font(size=16, bold=True);
        ws['A1'].alignment = align
        ws.append(["工程名称:", self.order_info['工程名称'], "", "", "发货方式:", "自提", "", "", "", "", "下单日期 :",
                   self.order_info['日期'], "", "单号:", self.order_info['单号']])
        ws.merge_cells('B2:D2');
        ws.merge_cells('F2:J2');
        ws.merge_cells('L2:N2');
        ws.merge_cells('P2:U2')
        # ... (此处省略重复的表头和数据填充逻辑) ...
        self.apply_full_style(ws, border, align)

    def draw_outbound_sheet(self, ws, border, align):
        # ... (出库单绘制逻辑) ...
        self.apply_full_style(ws, border, align)

    def draw_parts_sheet(self, ws, border, align):
        # ... (配件单绘制逻辑) ...
        self.apply_full_style(ws, border, align)

    def apply_full_style(self, ws, border, align):
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = align


if __name__ == "__main__":
    root = tk.Tk()
    app = OrderApp(root)
    root.mainloop()